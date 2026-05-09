"""Cell + VBA extraction.

Reads xlsm via openpyxl (formula text + cached values + sheet states + named
ranges + CF/DV) and VBA via oletools.olevba. Robust: per-cell / per-module
exceptions are logged into `parse_errors`, never propagated.

The `sanitize` flag, when True, replaces every non-formula cell value with
the constant string `<redacted>`. Formula text is preserved (it's structure,
not data).
"""

from __future__ import annotations

import re
import warnings
from dataclasses import dataclass
from pathlib import Path

# Silence openpyxl warnings on a few of the test files
warnings.filterwarnings("ignore", message="Unknown extension is not supported")
warnings.filterwarnings("ignore", message="Workbook contains no default style")
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

import openpyxl  # noqa: E402
from oletools.olevba import VBA_Parser  # noqa: E402

from .formula_utils import _col_letter

REDACTED_PLACEHOLDER = "<redacted>"

# VBA external/COM keyword set
VBA_EXTERNAL_KEYWORDS = (
    "CreateObject", "GetObject", "Shell", "ADODB", "WScript",
    "Environ", "Application.Run", "DDEInitiate", "URLDownloadToFile",
    "WinHttpRequest", "InternetExplorer.Application",
)


@dataclass
class CellRow:
    sheet: str
    sheet_state: str
    row: int
    col: int
    ref: str
    value: str
    formula: str
    cached_value: str


@dataclass
class VbaSubFunction:
    kind: str
    name: str


@dataclass
class VbaModule:
    name: str
    type: str
    line_count: int
    sub_functions: list
    external_keywords: list
    range_literals: list
    has_on_error_resume_next: bool
    source_text: str = ""  # raw module text — used by vba_classify; NOT serialized to JSON


@dataclass
class NamedRange:
    name: str
    scope: str
    ref: str


def extract_cells(
    path: Path, parse_errors: list, sanitize: bool = False,
) -> tuple[list, dict, list, int]:
    """
    Returns (cell_rows, sheet_meta_map, named_ranges_list, cf_total).
    When sanitize=True, every non-formula cell value is replaced with the
    placeholder `<redacted>`. Formulas / structure / counts are preserved.
    """
    cell_rows: list = []
    sheet_meta: dict = {}
    cf_total = 0

    wb = openpyxl.load_workbook(str(path), keep_vba=True, data_only=False)
    try:
        wb_data = openpyxl.load_workbook(str(path), keep_vba=True, data_only=True)
    except Exception as e:
        parse_errors.append({"sheet": "", "ref": "",
                             "error": f"data_only pass failed: {type(e).__name__}: {e}"})
        wb_data = None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_data = wb_data[sheet_name] if wb_data is not None and sheet_name in wb_data.sheetnames else None
        state = getattr(ws, "sheet_state", "visible") or "visible"

        rows_used = 0
        cols_used = 0
        n_nonempty = 0
        n_formula = 0

        try:
            for row in ws.iter_rows():
                for cell in row:
                    try:
                        raw = cell.value
                        if raw is None:
                            continue
                        r = int(cell.row)
                        c = int(cell.column)
                        rows_used = max(rows_used, r)
                        cols_used = max(cols_used, c)

                        formula = ""
                        value_str = ""
                        cached_str = ""

                        if isinstance(raw, str) and raw.startswith("="):
                            formula = raw
                            n_formula += 1
                        elif cell.data_type == "f":
                            s_raw = str(raw)
                            formula = s_raw if s_raw.startswith("=") else "=" + s_raw
                            n_formula += 1
                        else:
                            value_str = REDACTED_PLACEHOLDER if sanitize else str(raw)

                        if formula and ws_data is not None:
                            try:
                                cached = ws_data.cell(row=r, column=c).value
                                if cached is not None:
                                    if sanitize:
                                        # Preserve error tokens (#REF! etc) so risk indicators still work,
                                        # but redact normal cached values.
                                        s_cached = str(cached)
                                        if any(tok in s_cached for tok in (
                                                "#REF!", "#NAME?", "#DIV/0!", "#VALUE!",
                                                "#N/A", "#NUM!", "#NULL!")):
                                            cached_str = s_cached
                                        else:
                                            cached_str = REDACTED_PLACEHOLDER
                                    else:
                                        cached_str = str(cached)
                            except Exception:
                                cached_str = ""

                        cell_rows.append(CellRow(
                            sheet=sheet_name, sheet_state=state,
                            row=r, col=c, ref=f"{_col_letter(c)}{r}",
                            value=value_str, formula=formula, cached_value=cached_str,
                        ))
                        n_nonempty += 1
                    except Exception as e:
                        parse_errors.append({
                            "sheet": sheet_name,
                            "ref": getattr(cell, "coordinate", "?"),
                            "error": f"{type(e).__name__}: {e}",
                        })
        except Exception as e:
            parse_errors.append({
                "sheet": sheet_name, "ref": "",
                "error": f"sheet-level: {type(e).__name__}: {e}",
            })

        cf_count = 0
        try:
            cf_rules = ws.conditional_formatting
            if hasattr(cf_rules, "_cf_rules"):
                for _rng, rules in cf_rules._cf_rules.items():
                    cf_count += len(rules)
        except Exception:
            cf_count = 0
        cf_total += cf_count

        dv_count = 0
        try:
            if ws.data_validations and getattr(ws.data_validations, "dataValidation", None):
                dv_count = len(ws.data_validations.dataValidation)
        except Exception:
            dv_count = 0

        sheet_meta[sheet_name] = {
            "state": state,
            "rows_used": rows_used,
            "cols_used": cols_used,
            "n_nonempty": n_nonempty,
            "n_formula": n_formula,
            "max_ref": f"{_col_letter(max(cols_used, 1))}{max(rows_used, 1)}" if n_nonempty else "",
            "cf_count": cf_count,
            "dv_count": dv_count,
        }

    named_ranges: list = []
    try:
        for name, defn in wb.defined_names.items():
            try:
                if defn.destinations:
                    ref = ", ".join(f"{t}!{r}" for t, r in defn.destinations)
                else:
                    ref = defn.value or ""
            except Exception:
                ref = getattr(defn, "value", "") or ""
            named_ranges.append(NamedRange(name=name, scope="workbook", ref=ref))
    except Exception:
        pass
    for ws in wb.worksheets:
        dn = getattr(ws, "defined_names", None)
        if dn is None:
            continue
        try:
            items = list(dn.items())
        except Exception:
            continue
        for name, defn in items:
            try:
                if defn.destinations:
                    ref = ", ".join(f"{t}!{r}" for t, r in defn.destinations)
                else:
                    ref = defn.value or ""
            except Exception:
                ref = getattr(defn, "value", "") or ""
            named_ranges.append(NamedRange(name=name, scope=ws.title, ref=ref))
    named_ranges.sort(key=lambda nr: (nr.name.lower(), nr.scope.lower()))

    return cell_rows, sheet_meta, named_ranges, cf_total


_RE_SUB_FUNC = re.compile(
    r"^\s*(?:Public|Private|Friend)?\s*(?:Static\s+)?(Sub|Function|Property\s+Get|Property\s+Let|Property\s+Set)\s+([A-Za-z_][\w]*)",
    re.IGNORECASE | re.MULTILINE,
)
_RE_RANGE_LITERAL = re.compile(
    r"""(?ix)
    \b(?:Range|Cells|Sheets?|Worksheets?)\s*\(\s*
    (?P<arg>"[^"]*"|'[^']*')
    \s*\)
    """
)
_RE_ON_ERROR_RESUME = re.compile(r"\bOn\s+Error\s+Resume\s+Next\b", re.IGNORECASE)


def extract_vba(path: Path, parse_errors: list) -> tuple[list, str]:
    """Returns (vba_modules, vba_text_blob).

    Note: VBA source code is *not* sanitized. It's code, not user data, and the
    audit needs it to detect smells and classify modules. If sanitize is needed
    over VBA too, that's a future feature; users with truly secret VBA should not
    distribute the audit at all.
    """
    modules: list = []
    text_parts: list = []
    try:
        parser = VBA_Parser(str(path))
    except Exception as e:
        parse_errors.append({"sheet": "<vba>", "ref": "", "error": f"VBA_Parser init: {type(e).__name__}: {e}"})
        return modules, ""

    try:
        if not parser.detect_vba_macros():
            parser.close()
            return modules, ""
    except Exception as e:
        parse_errors.append({"sheet": "<vba>", "ref": "", "error": f"detect_vba_macros: {type(e).__name__}: {e}"})
        try:
            parser.close()
        except Exception:
            pass
        return modules, ""

    try:
        for filename, stream_path, vba_filename, vba_code in parser.extract_macros():
            try:
                if vba_filename.lower().endswith(".bas"):
                    mtype = "standard"
                elif vba_filename.lower().endswith(".cls"):
                    mtype = "class_or_document"
                elif vba_filename.lower().endswith(".frm"):
                    mtype = "userform"
                else:
                    mtype = "unknown"

                code = vba_code or ""
                lines = code.splitlines()

                subs: list = []
                for m in _RE_SUB_FUNC.finditer(code):
                    kind_raw = m.group(1).strip()
                    if kind_raw.lower().startswith("property"):
                        kind = "Property"
                    else:
                        kind = kind_raw[0].upper() + kind_raw[1:].lower()
                    subs.append(VbaSubFunction(kind=kind, name=m.group(2)))
                subs.sort(key=lambda s: (s.kind, s.name))

                ext_kw = sorted({kw for kw in VBA_EXTERNAL_KEYWORDS if kw in code})

                ranges = sorted({m.group("arg").strip('"').strip("'") for m in _RE_RANGE_LITERAL.finditer(code)})

                has_oer = bool(_RE_ON_ERROR_RESUME.search(code))

                modules.append(VbaModule(
                    name=vba_filename,
                    type=mtype,
                    line_count=len(lines),
                    sub_functions=subs,
                    external_keywords=ext_kw,
                    range_literals=ranges,
                    has_on_error_resume_next=has_oer,
                    source_text=code,
                ))

                text_parts.append("=" * 64)
                text_parts.append(f"=== Module: {vba_filename} (type: {mtype}) ===")
                text_parts.append(f"=== Stream: {stream_path} ===")
                text_parts.append("=" * 64)
                text_parts.append(code.rstrip())
                text_parts.append("")
            except Exception as e:
                parse_errors.append({
                    "sheet": "<vba>", "ref": vba_filename,
                    "error": f"module-extract: {type(e).__name__}: {e}",
                })
    except Exception as e:
        parse_errors.append({"sheet": "<vba>", "ref": "", "error": f"extract_macros: {type(e).__name__}: {e}"})

    try:
        parser.close()
    except Exception:
        pass

    modules.sort(key=lambda m: m.name.lower())
    return modules, "\n".join(text_parts)
