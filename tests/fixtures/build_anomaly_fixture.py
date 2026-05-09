"""Build a tiny xlsm fixture with a deliberate magic-number anomaly.

Run once to (re)generate `anomaly_fixture.xlsm`. Run inside the package venv:
    .venv/bin/python tests/fixtures/build_anomaly_fixture.py

The fixture has:
- Sheet 'Pricing' with 30 rows of `=B<n> * 0.85` for n in 2..30
- Row 27 (D27) deliberately uses 0.82 instead of 0.85 -> single outlier
- Sheet 'Constants' with B2 = 0.07 (used as a pillar input)
- Sheet 'Tax' with 25 rows of `=Constants!$B$2 * <constant>` — uses different
  number-token shape (no plain `=R*N` collision with Pricing's discount column)
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

OUT = Path(__file__).resolve().parent / "anomaly_fixture.xlsm"


def build():
    wb = openpyxl.Workbook()

    # ---- Pricing sheet ----
    ws = wb.active
    ws.title = "Pricing"
    ws["A1"] = "qty"
    ws["B1"] = "unit_price"
    ws["C1"] = "subtotal"
    ws["D1"] = "discounted"
    for r in range(2, 31):
        ws.cell(row=r, column=1, value=r * 10)
        ws.cell(row=r, column=2, value=100.0)
        # Deliberately make subtotal a non-trivial pattern that is DIFFERENT
        # from the discount formula so they don't collapse into one cluster.
        ws.cell(row=r, column=3, value=f"=A{r}*B{r}+0.01")  # +0.01 keeps it distinct
        # Most rows use 0.85 discount; row 27 uses 0.82 (deliberate anomaly).
        # Keep all D-column formulas in one tight cluster (`=Cn*N`) so the
        # outlier really pops.
        if r == 27:
            ws.cell(row=r, column=4, value=f"=C{r}*0.82")
        else:
            ws.cell(row=r, column=4, value=f"=C{r}*0.85")

    # ---- Constants sheet (used as pillar input) ----
    ws2 = wb.create_sheet("Constants")
    ws2["A1"] = "name"
    ws2["B1"] = "value"
    ws2["A2"] = "tax_rate"
    ws2["B2"] = 0.07

    # ---- Tax sheet — 25 cells reference Constants!B2 (creates a pillar) ----
    ws3 = wb.create_sheet("Tax")
    for r in range(2, 27):
        # Use a unique formula shape (not `=R*N`) so it doesn't merge clusters
        # with the discount column above.
        ws3.cell(row=r, column=1, value=f"=Constants!$B$2+0.5")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    tmp = OUT.with_suffix(".xlsx")
    wb.save(tmp)
    tmp.rename(OUT)
    print(f"wrote: {OUT}")


if __name__ == "__main__":
    build()
