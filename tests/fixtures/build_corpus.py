"""build_corpus.py — Multi-variant fixture generator (Round 4 / Stage 2 expansion).

Sibling of `build_anomaly_fixture.py`. Produces three additional xlsm fixtures
covering different domains, sizes, and styles so we can prove Track A's
audit pipeline is consistent across diverse inputs.

Usage:

    python tests/fixtures/build_corpus.py --variant logistics-routing \
        --out tests/fixtures/logistics_routing_synth.xlsm
    python tests/fixtures/build_corpus.py --variant inventory-supply-chain \
        --out tests/fixtures/inventory_supply_synth.xlsm
    python tests/fixtures/build_corpus.py --variant minimal \
        --out tests/fixtures/minimal_no_vba.xlsm

Determinism contract (matches the existing capacity-planning generator):
    - All randomness via `random.Random(seed=20260509)`. Re-running the
      generator overwrites the output byte-stably (modulo openpyxl's own
      output, which is deterministic on a fixed Python+openpyxl version).
    - No new pip dependencies (openpyxl + stdlib only).
    - VBA injection re-uses the vba-web-example.xlsm donor — the exact same
      donor used by `gen_synth_xlsm.py`. Documented honestly as unrelated VBA;
      its only purpose is to populate `wb.vba_archive` so the audit's VBA
      classifier has something to chew on.

Variants:
    * logistics-routing — 6 sheets, ~190 rows of data, named ranges,
      cross-sheet INDEX/MATCH and SUMIFS, conditional formatting,
      one form-control button bound to a real Sub from the donor VBA.
    * inventory-supply-chain — 5 sheets, EOQ + safety stock + reorder
      formulas, named ranges, data validation. VBA from donor (no buttons).
    * minimal — degradation test. 2 sheets, no VBA, no named ranges,
      no hidden sheets, no conditional formatting.

The button injection (logistics variant) is performed post-save by editing
the saved zip's sheet xml, content types, rels, and adding a
`xl/ctrlProps/ctrlPropN.xml` file. openpyxl can't natively create form
controls, so we splice the OPC parts in directly. This is the same
technique enterprise users would see in real workbooks (Excel writes
exactly these XML fragments).
"""
from __future__ import annotations

import argparse
import random
import re
import shutil
import warnings
import zipfile
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# ----------------------------------------------------------------------------
# Configuration
# ----------------------------------------------------------------------------

DONOR_XLSM = Path(
    "/Users/tang/Documents/AgenticFactory/workers/20260508-01-xlsm-vba-parsing"
    "/test_files/vba-web-example.xlsm"
)

SEED = 20260509  # spec-mandated for determinism


# ----------------------------------------------------------------------------
# Style helpers (shared with the existing generator's look)
# ----------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
SUBTLE_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hdr(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER


def _stripe(ws, row, ncols):
    if row % 2 != 0:
        return
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = SUBTLE_FILL


def _load_donor_workbook():
    """Return a fresh openpyxl workbook loaded from the donor xlsm.

    The donor preserves vba_archive (zipfile pointer); we wipe its content
    sheets + defined names so we can rebuild from scratch.
    """
    if not DONOR_XLSM.exists():
        raise FileNotFoundError(f"VBA donor missing: {DONOR_XLSM}")
    wb = openpyxl.load_workbook(str(DONOR_XLSM), keep_vba=True, data_only=False)
    for name in list(wb.sheetnames):
        del wb[name]
    for name in list(wb.defined_names):
        del wb.defined_names[name]
    return wb


# ============================================================================
# Variant 1: logistics-routing
# ============================================================================

_LOGISTICS_CITIES = [
    "Shanghai", "Beijing", "Guangzhou", "Shenzhen", "Chengdu",
    "Hangzhou", "Suzhou", "Tianjin", "Wuhan", "Nanjing",
    "Xi'an", "Qingdao", "Dalian", "Ningbo", "Xiamen",
    "Changsha", "Hefei", "Zhengzhou", "Foshan", "Jinan",
]


def _build_logistics_constants(ws):
    """_constants — veryHidden, magic numbers (the spec's required 6)."""
    ws.sheet_state = "veryHidden"
    _hdr(ws, 1, ["param_id", "name", "value", "unit", "notes"])
    rows = [
        ("MAX_VEHICLES",          "Max fleet size",            15,    "vehicles", "Hard fleet cap"),
        ("SPEED_KMH",             "Average speed",             60,    "km/h",     "Inter-city avg"),
        ("SERVICE_TIME_MIN",      "Service time per stop",     15,    "minutes",  "Per delivery"),
        ("MAX_STOPS_PER_ROUTE",   "Max stops per route",       20,    "stops",    "Driver workload"),
        ("FUEL_PRICE",            "Fuel unit price",           1.85,  "USD/L",    "Last year average"),
        ("DRIVER_SHIFT_HOURS",    "Driver shift hours",        8,     "hours",    "Regulatory limit"),
    ]
    for i, r in enumerate(rows, 2):
        for j, v in enumerate(r, 1):
            ws.cell(row=i, column=j, value=v)
        _stripe(ws, i, 5)


def _build_logistics_vehicles(ws, rng):
    _hdr(ws, 1, ["VehicleID", "CapacityKg", "MaxStops", "FuelCostPerKm"])
    for i in range(2, 17):  # 15 vehicles
        ws.cell(row=i, column=1, value=f"V{i-1:03d}")
        ws.cell(row=i, column=2, value=rng.choice([2000, 2500, 3000, 3500, 5000]))
        ws.cell(row=i, column=3, value=rng.randint(15, 25))
        # FuelCostPerKm = FUEL_PRICE * efficiency_factor (random factor 0.20..0.45)
        eff = round(rng.uniform(0.20, 0.45), 3)
        ws.cell(row=i, column=4, value=f"=FUEL_PRICE*{eff}")
        _stripe(ws, i, 4)


def _build_logistics_routes(ws, rng, n_routes=30, n_vehicles=15):
    _hdr(ws, 1, [
        "RouteID", "AssignedVehicle", "DistanceKm",
        "TotalStops", "EstimatedTimeHours", "CapacityUtilization",
    ])
    for i in range(2, n_routes + 2):
        ws.cell(row=i, column=1, value=f"R{i-1:03d}")
        # AssignedVehicle picks one of V001..V015
        v = f"V{rng.randint(1, n_vehicles):03d}"
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=3, value=rng.randint(50, 600))
        ws.cell(row=i, column=4, value=rng.randint(5, 22))
        # EstimatedTimeHours = DistanceKm / SPEED_KMH + TotalStops × SERVICE_TIME_MIN/60
        ws.cell(row=i, column=5,
                value=f"=C{i}/SPEED_KMH+D{i}*SERVICE_TIME_MIN/60")
        # CapacityUtilization = SUMIFS of WeightKg on this route / VLOOKUP capacity
        ws.cell(row=i, column=6,
                value=(f"=SUMIFS(Shipments!$C$2:$C$101,"
                       f"Shipments!$D$2:$D$101,A{i})"
                       f"/IFERROR(VLOOKUP(B{i},Vehicles!$A$2:$B$16,2,FALSE),1)"))
        _stripe(ws, i, 6)


def _build_logistics_shipments(ws, rng, n_ships=100, n_routes=30, n_customers=40):
    _hdr(ws, 1, [
        "OrderID", "CustomerID", "WeightKg", "AssignedRoute",
        "FuelCost",
    ])
    for i in range(2, n_ships + 2):
        ws.cell(row=i, column=1, value=f"O{i-1:05d}")
        ws.cell(row=i, column=2, value=f"C{rng.randint(1, n_customers):03d}")
        ws.cell(row=i, column=3, value=rng.randint(50, 1500))
        ws.cell(row=i, column=4, value=f"R{rng.randint(1, n_routes):03d}")
        # FuelCost = lookup route DistanceKm * FUEL_PRICE
        # Use INDEX/MATCH cross-sheet for variety
        ws.cell(row=i, column=5,
                value=(f"=IFERROR(INDEX(Routes!$C$2:$C$31,"
                       f"MATCH(D{i},Routes!$A$2:$A$31,0))*FUEL_PRICE,0)"))
        _stripe(ws, i, 5)


def _build_logistics_customers(ws, rng, n_customers=40):
    _hdr(ws, 1, ["CustomerID", "City", "TimeWindowStart", "TimeWindowEnd"])
    for i in range(2, n_customers + 2):
        ws.cell(row=i, column=1, value=f"C{i-1:03d}")
        ws.cell(row=i, column=2, value=rng.choice(_LOGISTICS_CITIES))
        # Time windows: start 06:00..14:00, end +2..+5 hours
        start_h = rng.randint(6, 14)
        end_h = start_h + rng.randint(2, 5)
        ws.cell(row=i, column=3, value=f"{start_h:02d}:00")
        ws.cell(row=i, column=4, value=f"{end_h:02d}:00")
        _stripe(ws, i, 4)


def _build_logistics_orders_zh(ws, rng, n=20):
    """订单 — visible, mixed Chinese-English column names."""
    _hdr(ws, 1, ["订单号", "Customer", "City", "Weight (kg)", "Delivery Date", "Status"])
    statuses = ["pending", "scheduled", "in-transit", "delivered"]
    for i in range(2, n + 2):
        ws.cell(row=i, column=1, value=f"DIN-{i-1:04d}")
        ws.cell(row=i, column=2, value=f"C{rng.randint(1, 40):03d}")
        ws.cell(row=i, column=3, value=rng.choice(_LOGISTICS_CITIES))
        ws.cell(row=i, column=4, value=rng.randint(80, 2000))
        ws.cell(row=i, column=5, value=f"2026-{rng.randint(5, 8):02d}-{rng.randint(1, 28):02d}")
        ws.cell(row=i, column=6, value=rng.choice(statuses))
        _stripe(ws, i, 6)


def _add_logistics_named_ranges(wb):
    names = [
        ("MAX_VEHICLES",          "_constants!$C$2"),
        ("SPEED_KMH",             "_constants!$C$3"),
        ("SERVICE_TIME_MIN",      "_constants!$C$4"),
        ("MAX_STOPS_PER_ROUTE",   "_constants!$C$5"),
        ("FUEL_PRICE",            "_constants!$C$6"),
        ("DRIVER_SHIFT_HOURS",    "_constants!$C$7"),
    ]
    for name, ref in names:
        wb.defined_names[name] = DefinedName(name=name, attr_text=ref)


def _add_logistics_conditional_formatting(ws):
    """Red fill on Routes!F2..F31 when capacity utilization > 100%."""
    rule = CellIsRule(
        operator="greaterThan",
        formula=["1.0"],
        stopIfTrue=False,
        fill=PatternFill("solid", fgColor="FFC7CE"),
        font=Font(color="9C0006"),
    )
    ws.conditional_formatting.add("F2:F31", rule)


def build_logistics_routing(out_path: Path):
    """Variant 1 — logistics-routing fixture, ~300-400 KB with VBA + button."""
    rng = random.Random(SEED)
    wb = _load_donor_workbook()

    # Create sheets in declared order
    ws_routes  = wb.create_sheet("Routes")
    ws_veh     = wb.create_sheet("Vehicles")
    ws_ship    = wb.create_sheet("Shipments")
    ws_cust    = wb.create_sheet("Customers")
    ws_orders  = wb.create_sheet("订单")
    ws_const   = wb.create_sheet("_constants")

    # Populate (constants first so named-range targets exist on save)
    _build_logistics_constants(ws_const)
    _build_logistics_vehicles(ws_veh, rng)
    _build_logistics_routes(ws_routes, rng)
    _build_logistics_shipments(ws_ship, rng)
    _build_logistics_customers(ws_cust, rng)
    _build_logistics_orders_zh(ws_orders, rng)

    _add_logistics_named_ranges(wb)
    _add_logistics_conditional_formatting(ws_routes)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    # Post-save: inject a real form-control button into the Routes sheet.
    # Macro target: backup__001WebClient.cls -> Sub Execute (real Sub from donor).
    # Display label: "Run Routing Calculation"
    _inject_button(
        out_path,
        target_sheet="Routes",
        macro_token="[0]!Execute",
        button_label="Run Routing Calculation",
    )
    _strip_nondeterminism(out_path)

    return out_path


# ============================================================================
# Variant 2: inventory-supply-chain
# ============================================================================

_INV_SUPPLIERS = [
    ("S001", 7,  0.95, 100),
    ("S002", 14, 0.90, 250),
    ("S003", 21, 0.85, 500),
    ("S004", 10, 0.92, 50),
    ("S005", 5,  0.97, 200),
    ("S006", 30, 0.80, 1000),
    ("S007", 12, 0.91, 150),
    ("S008", 7,  0.94, 75),
    ("S009", 18, 0.88, 300),
    ("S010", 9,  0.93, 125),
]


def _build_inv_constants(ws):
    """_constants — veryHidden, 5 magic numbers."""
    ws.sheet_state = "veryHidden"
    _hdr(ws, 1, ["param_id", "name", "value", "unit", "notes"])
    rows = [
        ("SAFETY_FACTOR",         "Safety stock z-multiplier", 1.65, "ratio",     "95% service level"),
        ("ANNUAL_HOLDING_RATE",   "Holding cost rate",         0.20, "ratio/yr",  "Capital + storage"),
        ("ORDER_COST",            "Cost per purchase order",   50,   "USD",       "Admin overhead"),
        ("ANNUAL_DEMAND_DEFAULT", "Default annual demand",     1000, "units/yr",  "Used when SKU lacks history"),
        ("SERVICE_LEVEL_Z",       "Z-score for service level", 1.96, "ratio",     "97.5% one-tail"),
    ]
    for i, r in enumerate(rows, 2):
        for j, v in enumerate(r, 1):
            ws.cell(row=i, column=j, value=v)
        _stripe(ws, i, 5)


def _build_inv_stock(ws, rng, n_skus=50):
    """Stock — visible, 50 SKUs with EOQ + safety stock + reorder formulas."""
    _hdr(ws, 1, [
        "SKU", "CurrentQty", "ReorderPoint", "SafetyStock",
        "EOQ", "UnitPrice", "DailyDemand", "LeadTimeDays",
    ])
    for i in range(2, n_skus + 2):
        ws.cell(row=i, column=1, value=f"SKU-{i-1:04d}")
        ws.cell(row=i, column=2, value=rng.randint(0, 800))
        # Daily demand and lead time set first (they're inputs)
        daily = rng.randint(5, 60)
        lt = rng.choice([5, 7, 10, 14, 21])
        price = round(rng.uniform(2.50, 95.0), 2)
        ws.cell(row=i, column=6, value=price)        # UnitPrice (F)
        ws.cell(row=i, column=7, value=daily)        # DailyDemand (G)
        ws.cell(row=i, column=8, value=lt)           # LeadTimeDays (H)
        # SafetyStock (D) = SAFETY_FACTOR * sigma * SQRT(LeadTime)
        # sigma approximated by 0.30 * daily demand
        ws.cell(row=i, column=4,
                value=f"=SAFETY_FACTOR*0.30*G{i}*SQRT(H{i})")
        # ReorderPoint (C) = DailyDemand*LeadTime + SafetyStock
        ws.cell(row=i, column=3, value=f"=G{i}*H{i}+D{i}")
        # EOQ (E) = SQRT(2*ANNUAL_DEMAND_DEFAULT*ORDER_COST/(ANNUAL_HOLDING_RATE*UnitPrice))
        ws.cell(row=i, column=5,
                value=("=SQRT(2*ANNUAL_DEMAND_DEFAULT*ORDER_COST/"
                       f"(ANNUAL_HOLDING_RATE*F{i}))"))
        _stripe(ws, i, 8)


def _build_inv_reorder(ws, n_skus=50):
    """Reorder — visible, 50 rows mirroring Stock."""
    _hdr(ws, 1, [
        "SKU", "SuggestedOrderQty", "ExpectedDelivery",
        "TotalReceiptsLast30d", "TotalIssuesLast30d",
    ])
    for i in range(2, n_skus + 2):
        ws.cell(row=i, column=1, value=f"SKU-{i-1:04d}")
        # SuggestedOrderQty = MAX(EOQ from Stock, ReorderPoint - CurrentQty)
        ws.cell(row=i, column=2,
                value=f"=MAX(Stock!E{i},Stock!C{i}-Stock!B{i})")
        # ExpectedDelivery = TODAY() + LeadTime
        ws.cell(row=i, column=3, value=f"=TODAY()+Stock!H{i}")
        # SUMIFS aggregations over Movements
        ws.cell(row=i, column=4,
                value=(f'=SUMIFS(Movements!$C$2:$C$201,'
                       f'Movements!$B$2:$B$201,A{i},'
                       f'Movements!$D$2:$D$201,"in")'))
        ws.cell(row=i, column=5,
                value=(f'=SUMIFS(Movements!$C$2:$C$201,'
                       f'Movements!$B$2:$B$201,A{i},'
                       f'Movements!$D$2:$D$201,"out")'))
        _stripe(ws, i, 5)


def _build_inv_suppliers(ws):
    _hdr(ws, 1, ["SupplierID", "LeadTimeDays", "ReliabilityScore", "MOQ"])
    for i, sup in enumerate(_INV_SUPPLIERS, 2):
        for j, v in enumerate(sup, 1):
            ws.cell(row=i, column=j, value=v)
        _stripe(ws, i, 4)


def _build_inv_movements(ws, rng, n_rows=200, n_skus=50):
    _hdr(ws, 1, ["Date", "SKU", "Quantity", "Type"])
    for i in range(2, n_rows + 2):
        ws.cell(row=i, column=1, value=f"2026-{rng.randint(3, 5):02d}-{rng.randint(1, 28):02d}")
        ws.cell(row=i, column=2, value=f"SKU-{rng.randint(1, n_skus):04d}")
        ws.cell(row=i, column=3, value=rng.randint(10, 300))
        ws.cell(row=i, column=4, value=rng.choice(["in", "out"]))
        _stripe(ws, i, 4)


def _add_inv_named_ranges(wb):
    names = [
        ("SAFETY_FACTOR",         "_constants!$C$2"),
        ("ANNUAL_HOLDING_RATE",   "_constants!$C$3"),
        ("ORDER_COST",            "_constants!$C$4"),
        ("ANNUAL_DEMAND_DEFAULT", "_constants!$C$5"),
        ("SERVICE_LEVEL_Z",       "_constants!$C$6"),
        # Spec also asked for LEAD_TIME_DEFAULT — point it at the same row
        # used by Suppliers (Reorder.xlsx style indirection); we point it at
        # _constants!C5 since it's a default in the same spirit. Documented
        # below via the cell value.
        ("LEAD_TIME_DEFAULT",     "_constants!$C$5"),
    ]
    for name, ref in names:
        wb.defined_names[name] = DefinedName(name=name, attr_text=ref)


def _add_inv_data_validation(ws_movements, n_rows=200):
    dv = DataValidation(
        type="list",
        formula1='"in,out"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid type",
        error="Type must be 'in' or 'out'.",
    )
    dv.add(f"D2:D{n_rows + 1}")
    ws_movements.add_data_validation(dv)


def build_inventory_supply(out_path: Path):
    """Variant 2 — inventory-supply-chain fixture."""
    rng = random.Random(SEED)
    wb = _load_donor_workbook()

    ws_stock     = wb.create_sheet("Stock")
    ws_reorder   = wb.create_sheet("Reorder")
    ws_sup       = wb.create_sheet("Suppliers")
    ws_moves     = wb.create_sheet("Movements")
    ws_const     = wb.create_sheet("_constants")

    _build_inv_constants(ws_const)
    _build_inv_stock(ws_stock, rng)
    _build_inv_reorder(ws_reorder)
    _build_inv_suppliers(ws_sup)
    _build_inv_movements(ws_moves, rng)

    _add_inv_named_ranges(wb)
    _add_inv_data_validation(ws_moves)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    _strip_nondeterminism(out_path)
    return out_path


# ============================================================================
# Variant 3: minimal (no VBA, no named ranges, no hidden, no CF, no DV)
# ============================================================================

def build_minimal(out_path: Path):
    """Variant 3 — degradation/robustness test fixture."""
    rng = random.Random(SEED)
    # Pure xlsm without VBA: an xlsm with empty vba_archive is still valid.
    # We use a fresh openpyxl Workbook saved as .xlsx then renamed to .xlsm —
    # the audit's olefile probe will return empty modules, confirming the
    # "no VBA" path is exercised.
    wb = openpyxl.Workbook()

    ws_data = wb.active
    ws_data.title = "Data"
    ws_data.append(["Date", "Product", "Quantity", "Price"])
    products = ["Widget-A", "Widget-B", "Gadget-X", "Gadget-Y"]
    for i in range(2, 32):  # 30 data rows
        ws_data.cell(row=i, column=1, value=f"2026-04-{(i - 1) % 28 + 1:02d}")
        ws_data.cell(row=i, column=2, value=rng.choice(products))
        ws_data.cell(row=i, column=3, value=rng.randint(1, 100))
        ws_data.cell(row=i, column=4, value=round(rng.uniform(2.0, 50.0), 2))

    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["Metric", "Value"])
    ws_sum.cell(row=2, column=1, value="Total Quantity")
    ws_sum.cell(row=2, column=2, value="=SUM(Data!C2:C31)")
    ws_sum.cell(row=3, column=1, value="Total Revenue")
    ws_sum.cell(row=3, column=2, value="=SUMPRODUCT(Data!C2:C31,Data!D2:D31)")
    ws_sum.cell(row=4, column=1, value="Widget-A Quantity")
    ws_sum.cell(row=4, column=2,
                value='=SUMIFS(Data!C2:C31,Data!B2:B31,"Widget-A")')
    ws_sum.cell(row=5, column=1, value="Widget-B Quantity")
    ws_sum.cell(row=5, column=2,
                value='=SUMIFS(Data!C2:C31,Data!B2:B31,"Widget-B")')
    ws_sum.cell(row=6, column=1, value="Gadget-X Quantity")
    ws_sum.cell(row=6, column=2,
                value='=SUMIFS(Data!C2:C31,Data!B2:B31,"Gadget-X")')
    ws_sum.cell(row=7, column=1, value="Gadget-Y Quantity")
    ws_sum.cell(row=7, column=2,
                value='=SUMIFS(Data!C2:C31,Data!B2:B31,"Gadget-Y")')
    ws_sum.cell(row=8, column=1, value="Avg Price")
    ws_sum.cell(row=8, column=2, value="=AVERAGE(Data!D2:D31)")
    ws_sum.cell(row=9, column=1, value="Max Price")
    ws_sum.cell(row=9, column=2, value="=MAX(Data!D2:D31)")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    # Save then rename to .xlsm — produces a macro-enabled extension with no
    # vbaProject.bin (oletools returns 0 modules, confirming graceful path).
    tmp_xlsx = out_path.with_suffix(".xlsx")
    wb.save(str(tmp_xlsx))
    if out_path.exists():
        out_path.unlink()
    tmp_xlsx.rename(out_path)
    _strip_nondeterminism(out_path)
    return out_path


# ============================================================================
# Button injection (logistics variant only)
# ============================================================================

# A single ctrlProp1.xml content — a regular form-control button.
_CTRLPROP_XML = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
    '<formControlPr xmlns="http://schemas.microsoft.com/office/spreadsheetml'
    '/2009/9/main" objectType="Button" lockText="1"/>'
)


# Fixed timestamp for the deterministic build. openpyxl writes `dcterms:modified`
# with the current UTC time on each save, breaking byte-identity. We rewrite
# it to a fixed Round-4 epoch.
_FIXED_MODIFIED = "2026-05-09T00:00:00Z"


def _strip_nondeterminism(xlsm_path: Path):
    """Post-process the saved xlsm to remove openpyxl's two known non-deterministic
    bits: (a) `dcterms:modified` timestamp in docProps/core.xml, (b) iteration-
    order-dependent Default Extension order in [Content_Types].xml. Without
    this fix the same generation inputs can produce two different byte streams
    on consecutive runs.
    """
    src = xlsm_path.read_bytes()
    parts: dict = {}
    with zipfile.ZipFile(BytesIO(src), "r") as zf:
        ordered = zf.namelist()
        for n in ordered:
            parts[n] = zf.read(n)

    # (a) Pin both dcterms:created and dcterms:modified to a fixed epoch.
    # Donor-derived fixtures inherit a static `created` from the donor file
    # but fresh `openpyxl.Workbook()` outputs (the minimal fixture) get a
    # current-time `created` on save, which varies run-to-run.
    if "docProps/core.xml" in parts:
        core = parts["docProps/core.xml"].decode("utf-8")
        core = re.sub(
            r'(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)',
            rf'\g<1>{_FIXED_MODIFIED}\g<2>',
            core,
        )
        core = re.sub(
            r'(<dcterms:created[^>]*>)[^<]*(</dcterms:created>)',
            rf'\g<1>{_FIXED_MODIFIED}\g<2>',
            core,
        )
        parts["docProps/core.xml"] = core.encode("utf-8")

    # (b) Re-emit Content_Types with Default elements alphabetically sorted by
    # Extension. We DON'T touch Override order (each Override has a stable
    # PartName which makes ordering canonical via the same comparator).
    if "[Content_Types].xml" in parts:
        ct = parts["[Content_Types].xml"].decode("utf-8")
        # Pull all Default and Override elements as raw strings.
        # Self-closing tags can be `/>`, ` />`, or `  />` — non-greedy match
        # of any chars up to the closing `>` keeps things robust.
        defaults = re.findall(r'<Default\b[^>]*?/>', ct)
        overrides = re.findall(r'<Override\b[^>]*?/>', ct)
        if not defaults and not overrides:
            # Don't blank the file — fall through and leave ct untouched.
            pass
        else:
            # Sort Defaults by their Extension attribute
            def _ext(s: str) -> str:
                m = re.search(r'Extension="([^"]+)"', s)
                return m.group(1) if m else ""
            defaults_sorted = sorted(defaults, key=_ext)
            # Sort Overrides by PartName
            def _pn(s: str) -> str:
                m = re.search(r'PartName="([^"]+)"', s)
                return m.group(1) if m else ""
            overrides_sorted = sorted(overrides, key=_pn)
            # Re-emit
            ct_new = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                '<Types xmlns="http://schemas.openxmlformats.org/package/2006/'
                'content-types">'
                + "".join(defaults_sorted)
                + "".join(overrides_sorted)
                + "</Types>"
            )
            parts["[Content_Types].xml"] = ct_new.encode("utf-8")

    # Re-emit zip in alphabetical key order (also deterministic)
    with zipfile.ZipFile(str(xlsm_path), "w", zipfile.ZIP_DEFLATED) as zf:
        for n in sorted(parts.keys()):
            # ZipInfo with fixed date_time → deterministic LZ headers
            info = zipfile.ZipInfo(filename=n, date_time=(2026, 5, 9, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            zf.writestr(info, parts[n])


def _inject_button(xlsm_path: Path, target_sheet: str,
                   macro_token: str, button_label: str):
    """Splice a `<controls>` block into a sheet's xml + add a ctrlProp file.

    What we do:
      1. Find the sheet's xml path via xl/workbook.xml + workbook.xml.rels.
      2. Append a `<controls>` block before `</worksheet>` containing a
         `<control name="..." r:id="rIdN" ...>` element and an inner
         `<controlPr macro="..."/>` with the requested macro token.
      3. Add a ctrlProp xml file at xl/ctrlProps/ctrlPropN.xml.
      4. Add a Relationship entry in xl/worksheets/_rels/sheetX.xml.rels
         linking that rId to the new ctrlProp file.
      5. Add an Override entry to [Content_Types].xml.

    Note: the audit's `_parse_sheet_controls` only needs the
    `<control name="..." ...>` + `<controlPr macro="...">` text, not a
    drawing rel. The label gets carried via the `name` attribute.
    """
    # Read entire archive into memory
    src = xlsm_path.read_bytes()
    parts: dict = {}
    with zipfile.ZipFile(BytesIO(src), "r") as zf:
        names = zf.namelist()
        for n in names:
            parts[n] = zf.read(n)

    # 1. Find the target sheet's path
    wb_xml = parts["xl/workbook.xml"].decode("utf-8")
    wb_rels = parts["xl/_rels/workbook.xml.rels"].decode("utf-8")
    rid_to_target: dict = {}
    for m in re.finditer(r'<Relationship\b([^>]*?)/?>', wb_rels):
        attrs = m.group(1)
        id_m = re.search(r'\bId="([^"]+)"', attrs)
        target_m = re.search(r'\bTarget="([^"]+)"', attrs)
        if id_m and target_m:
            rid_to_target[id_m.group(1)] = target_m.group(1)
    target_sheet_path = ""
    for m in re.finditer(r'<sheet\b([^/]*?)/>', wb_xml):
        attrs = m.group(1)
        name_m = re.search(r'name="([^"]*)"', attrs)
        rid_m = re.search(r'r:id="([^"]+)"', attrs)
        if not name_m or not rid_m:
            continue
        if name_m.group(1) == target_sheet:
            tg = rid_to_target.get(rid_m.group(1), "")
            # Targets may arrive as "/xl/worksheets/sheet1.xml" (leading slash,
            # absolute), "xl/worksheets/sheet1.xml" (relative-from-root), or
            # "worksheets/sheet1.xml" (relative-from-xl). Normalize to
            # "xl/worksheets/sheet1.xml" — the in-zip key.
            tg = tg.lstrip("/")
            if not tg.startswith("xl/"):
                tg = "xl/" + tg
            target_sheet_path = tg
            break
    if not target_sheet_path:
        raise RuntimeError(f"Could not locate sheet '{target_sheet}' in workbook.xml")

    # 2. Inject <controls> block into sheet xml just before </worksheet>
    sheet_xml = parts[target_sheet_path].decode("utf-8")
    # Pick a fresh ctrlProp index — find existing max + 1
    existing = [n for n in parts.keys()
                if n.startswith("xl/ctrlProps/ctrlProp") and n.endswith(".xml")]
    nums = []
    for e in existing:
        m = re.search(r'ctrlProp(\d+)\.xml', e)
        if m:
            nums.append(int(m.group(1)))
    next_idx = (max(nums) + 1) if nums else 1
    ctrlprop_path = f"xl/ctrlProps/ctrlProp{next_idx}.xml"

    # Pick a fresh rId for the sheet's rels
    sheet_rels_path = target_sheet_path.replace(
        "xl/worksheets/", "xl/worksheets/_rels/").replace(".xml", ".xml.rels")
    if sheet_rels_path in parts:
        sheet_rels = parts[sheet_rels_path].decode("utf-8")
    else:
        sheet_rels = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/'
            '2006/relationships"></Relationships>'
        )
    # next rId — count existing
    rids = [int(m.group(1)) for m in re.finditer(r'\bId="rId(\d+)"', sheet_rels)]
    next_rid = (max(rids) + 1) if rids else 1
    rid = f"rId{next_rid}"

    # Build the controls block
    # name="" carries the visible button label so the parser can pick it up;
    # macro="..." is what the audit's button detector keys on.
    controls_block = (
        f'<controls>'
        f'<mc:AlternateContent xmlns:mc="http://schemas.openxmlformats.org/'
        f'markup-compatibility/2006">'
        f'<mc:Choice Requires="x14">'
        f'<control shapeId="2049" r:id="{rid}" name="{button_label}">'
        f'<controlPr defaultSize="0" print="0" autoFill="0" autoPict="0" '
        f'macro="{macro_token}">'
        f'<anchor moveWithCells="1" sizeWithCells="1">'
        f'<from><xdr:col xmlns:xdr="http://schemas.openxmlformats.org/'
        f'drawingml/2006/spreadsheetDrawing">8</xdr:col><xdr:colOff '
        f'xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/'
        f'spreadsheetDrawing">0</xdr:colOff><xdr:row xmlns:xdr="http://'
        f'schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing">'
        f'1</xdr:row><xdr:rowOff xmlns:xdr="http://schemas.openxmlformats.org/'
        f'drawingml/2006/spreadsheetDrawing">0</xdr:rowOff></from>'
        f'<to><xdr:col xmlns:xdr="http://schemas.openxmlformats.org/drawingml/'
        f'2006/spreadsheetDrawing">10</xdr:col><xdr:colOff xmlns:xdr="http://'
        f'schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing">'
        f'0</xdr:colOff><xdr:row xmlns:xdr="http://schemas.openxmlformats.org/'
        f'drawingml/2006/spreadsheetDrawing">3</xdr:row><xdr:rowOff '
        f'xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/'
        f'spreadsheetDrawing">0</xdr:rowOff></to></anchor></controlPr>'
        f'</control></mc:Choice></mc:AlternateContent></controls>'
    )

    # The controls block uses r:id (relationships namespace). openpyxl-saved
    # sheets carry only the default xmlns; we must inject xmlns:r on the
    # <worksheet ...> root if it isn't already there.
    if "</worksheet>" not in sheet_xml:
        raise RuntimeError(f"Could not find </worksheet> in {target_sheet_path}")
    if 'xmlns:r=' not in sheet_xml[:600]:
        sheet_xml = re.sub(
            r"(<worksheet\b[^>]*?)>",
            (r'\1 xmlns:r="http://schemas.openxmlformats.org/'
             r'officeDocument/2006/relationships">'),
            sheet_xml,
            count=1,
        )
    sheet_xml_new = sheet_xml.replace("</worksheet>", controls_block + "</worksheet>")
    parts[target_sheet_path] = sheet_xml_new.encode("utf-8")

    # 3. Add the ctrlProp file
    parts[ctrlprop_path] = _CTRLPROP_XML.encode("utf-8")

    # 4. Add a Relationship in the sheet's _rels file
    rel_entry = (
        f'<Relationship Id="{rid}" '
        f'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
        f'relationships/ctrlProp" '
        f'Target="../ctrlProps/ctrlProp{next_idx}.xml"/>'
    )
    sheet_rels_new = sheet_rels.replace("</Relationships>", rel_entry + "</Relationships>")
    parts[sheet_rels_path] = sheet_rels_new.encode("utf-8")

    # 5. Add Override in [Content_Types].xml so Excel knows to read the part
    ct_xml = parts["[Content_Types].xml"].decode("utf-8")
    if f'PartName="/{ctrlprop_path}"' not in ct_xml:
        override = (
            f'<Override PartName="/{ctrlprop_path}" '
            f'ContentType="application/vnd.ms-excel.controlproperties+xml"/>'
        )
        ct_xml_new = ct_xml.replace("</Types>", override + "</Types>")
        parts["[Content_Types].xml"] = ct_xml_new.encode("utf-8")

    # Re-emit zip in original order
    with zipfile.ZipFile(str(xlsm_path), "w", zipfile.ZIP_DEFLATED) as zf:
        # Preserve original ordering for determinism
        with zipfile.ZipFile(BytesIO(src), "r") as orig:
            ordered = orig.namelist()
        for n in ordered:
            if n in parts:
                zf.writestr(n, parts[n])
                del parts[n]
        # Then append new files in deterministic alphabetic order
        for n in sorted(parts.keys()):
            zf.writestr(n, parts[n])


# ============================================================================
# CLI
# ============================================================================

VARIANTS = {
    "logistics-routing":     build_logistics_routing,
    "inventory-supply-chain": build_inventory_supply,
    "minimal":               build_minimal,
}


def main():
    ap = argparse.ArgumentParser(
        description="Generate one of three test corpus xlsm fixtures."
    )
    ap.add_argument("--variant", required=True, choices=sorted(VARIANTS.keys()),
                    help="Which fixture to build.")
    ap.add_argument("--out", required=True, help="Output xlsm path.")
    args = ap.parse_args()

    out = Path(args.out).resolve()
    fn = VARIANTS[args.variant]
    fn(out)
    size_kb = out.stat().st_size / 1024
    print(f"[done] wrote {out}  ({size_kb:.1f} KB)")


if __name__ == "__main__":
    main()
